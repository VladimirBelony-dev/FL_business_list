import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

class CorpsFormatter:
    """Apply consistent Corps formatting to files with all columns preserved"""
    
    def __init__(self):
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Corps standard styling - matching your spreadsheet
        self.light_blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11, name='Calibri')
        self.cell_font = Font(size=11, name='Calibri')
        self.thin_border = Border(
            left=Side(style='thin', color='B4C6E7'),
            right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'),
            bottom=Side(style='thin', color='B4C6E7')
        )
    
    def clean_address_data(self, df):
        """Clean address data to be single-line"""
        if 'Address' in df.columns:
            print("  Cleaning address data for single-line display...")
            df['Address'] = df['Address'].astype(str).str.replace('\n', ' ').str.replace('\r', ' ')
            df['Address'] = df['Address'].str.replace(r'\s+', ' ', regex=True)  # Multiple spaces to single space
            df['Address'] = df['Address'].str.strip()
            print("  ✓ Address data cleaned")
        
        if 'City State Zip' in df.columns:
            print("  Cleaning City State Zip data...")
            df['City State Zip'] = df['City State Zip'].astype(str).str.replace('\n', ' ').str.replace('\r', ' ')
            df['City State Zip'] = df['City State Zip'].str.replace(r'\s+', ' ', regex=True)
            df['City State Zip'] = df['City State Zip'].str.strip()
            print("  ✓ City State Zip data cleaned")
        
        return df
    
    def load_complete_file(self, filepath):
        """Load the file with all columns"""
        print(f"Loading {filepath}...")
        
        try:
            # Load all sheets to understand structure
            xl_file = pd.ExcelFile(filepath)
            print(f"  Found sheets: {', '.join(xl_file.sheet_names)}")
            
            # Load the main data sheet (usually first or 'Full Data')
            if 'Full Data' in xl_file.sheet_names:
                df = pd.read_excel(filepath, sheet_name='Full Data')
                print(f"  ✓ Loaded 'Full Data' sheet with {len(df):,} rows and {len(df.columns)} columns")
            else:
                df = pd.read_excel(filepath, sheet_name=0)
                print(f"  ✓ Loaded first sheet with {len(df):,} rows and {len(df.columns)} columns")
            
            print(f"  Columns found: {', '.join(df.columns.tolist())}")
            
            # Clean address data
            df = self.clean_address_data(df)
            
            return df
            
        except Exception as e:
            print(f"  ✗ Error: {e}")
            return None
    
    def apply_corps_formatting_to_all_columns(self, worksheet):
        """Apply Corps formatting to ALL columns in the worksheet"""
        
        print(f"Applying Corps formatting to {worksheet.max_column} columns...")
        
        # Auto-adjust column widths based on content
        for col_num in range(1, worksheet.max_column + 1):
            col_letter = worksheet.cell(row=1, column=col_num).column_letter
            
            # Set minimum and maximum widths
            max_length = 0
            
            # Check header
            header_length = len(str(worksheet.cell(row=1, column=col_num).value or ''))
            max_length = max(max_length, header_length)
            
            # Sample first 100 rows for width
            for row_num in range(2, min(102, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with min/max limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Format header row (all columns)
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self.thin_border
        
        # Apply alternating row colors to ALL columns
        print(f"  Applying alternating row colors to {worksheet.max_row - 1} data rows...")
        for row_num in range(2, worksheet.max_row + 1):
            # Alternating blue/white rows
            fill = self.light_blue_fill if row_num % 2 == 0 else self.white_fill
            
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = fill
                cell.font = self.cell_font
                cell.border = self.thin_border
                
                # Special handling for Address column - no text wrapping, single line
                # Check if this is the Address column by looking at the header
                header_value = str(worksheet.cell(row=1, column=col_num).value or '').lower()
                if 'address' in header_value:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        print(f"  ✓ Applied alternating colors: {self.light_blue_fill.start_color} and {self.white_fill.start_color}")
        
        # Set row heights
        worksheet.row_dimensions[1].height = 30  # Header row
        for row_num in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_num].height = 20
        
        print("  ✓ Formatting applied to all columns")
    
    def create_formatted_output(self, df, output_filename=None):
        """Create output with consistent formatting across all columns"""
        
        if output_filename is None:
            output_filename = f'Corps_10-2-25_FORMATTED_{self.timestamp}.xlsx'
        
        print(f"\nCreating formatted output: {output_filename}")
        
        # Save to Excel
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Main sheet with ALL columns
            df.to_excel(writer, sheet_name='Full Data', index=False)
            
            # Create Corps view (4 main columns only)
            corps_columns = ['Company', 'Officer', 'Address', 'City State Zip']
            corps_view = df[corps_columns].copy() if all(col in df.columns for col in corps_columns) else df.iloc[:, :4]
            corps_view.to_excel(writer, sheet_name='Corps View', index=False)
            
            # Statistics
            stats_data = []
            stats_data.append(['Total Rows', f"{len(df):,}"])
            stats_data.append(['Total Columns', f"{len(df.columns)}"])
            
            # Count filled data for key columns if they exist
            if 'Officer' in df.columns:
                filled_officer = len(df[df['Officer'].notna() & (df['Officer'] != '')])
                stats_data.append(['Rows with Officer Data', f"{filled_officer:,} ({filled_officer/len(df)*100:.1f}%)"])
            
            if 'Address' in df.columns:
                filled_address = len(df[df['Address'].notna() & (df['Address'] != '')])
                stats_data.append(['Rows with Address Data', f"{filled_address:,} ({filled_address/len(df)*100:.1f}%)"])
            
            stats_data.append(['Processing Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            
            stats_df = pd.DataFrame(stats_data, columns=['Metric', 'Value'])
            stats_df.to_excel(writer, sheet_name='Statistics', index=False)
        
        # Load and format the workbook
        wb = load_workbook(output_filename)
        
        # Format the Full Data sheet (all columns)
        ws_full = wb['Full Data']
        self.apply_corps_formatting_to_all_columns(ws_full)
        
        # Format the Corps View sheet
        ws_corps = wb['Corps View']
        self.apply_corps_formatting_to_all_columns(ws_corps)
        
        # Save formatted workbook
        wb.save(output_filename)
        
        print(f"  ✓ Saved: {output_filename}")
        
        return output_filename
    
    def process_files(self, file1, file2=None):
        """Process one or two files and apply consistent formatting"""
        
        print("="*80)
        print("CORPS DATA FORMATTER - PRESERVING ALL COLUMNS")
        print("="*80)
        print(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        # Load the primary file (with more columns)
        df1 = self.load_complete_file(file1)
        
        if df1 is None:
            print("Failed to load primary file!")
            return None
        
        # If second file provided, check if we need to merge
        if file2:
            df2 = self.load_complete_file(file2)
            
            if df2 is not None:
                print(f"\nComparing files...")
                print(f"  File 1: {len(df1.columns)} columns")
                print(f"  File 2: {len(df2.columns)} columns")
                
                # Use the file with more columns as base
                if len(df2.columns) > len(df1.columns):
                    print("  Using File 2 as base (more columns)")
                    df_final = df2
                else:
                    print("  Using File 1 as base (more columns)")
                    df_final = df1
            else:
                df_final = df1
        else:
            df_final = df1
        
        # Remove any completely empty rows
        df_final = df_final.dropna(how='all')
        
        # Sort by Company name if that column exists
        if 'Company' in df_final.columns:
            df_final = df_final.sort_values('Company', ignore_index=True)
        
        # Clean address data one more time to ensure single-line format
        df_final = self.clean_address_data(df_final)
        
        # Create formatted output
        output_file = self.create_formatted_output(df_final)
        
        # Print summary
        print("\n" + "="*80)
        print("FORMATTING COMPLETE!")
        print("="*80)
        print(f"Output file: {output_file}")
        print(f"Total rows: {len(df_final):,}")
        print(f"Total columns: {len(df_final.columns)}")
        print("\nThe file contains:")
        print("  • 'Full Data' - All columns with Corps blue/white formatting")
        print("  • 'Corps View' - Standard 4-column Corps view")
        print("  • 'Statistics' - Data summary")
        print("\nAll sheets have consistent Corps coloring and professional formatting.")
        
        return df_final

def main():
    """Main execution function"""
    
    formatter = CorpsFormatter()
    
    # Process your files
    # Since the first file has more columns, we'll use it as the primary
    result = formatter.process_files(
        file1='Corps_10-2-25_COMPLETE_20251004_180017.xlsx',  # Has more columns
        file2='Corps_10-2-25_COMPLETE_20251004_180317.xlsx'   # For comparison
    )
    
    if result is not None:
        print("\n✓ Successfully created formatted file with all columns preserved!")
        print("  The blue/white alternating row coloring is applied across ALL columns.")
    
    return result

if __name__ == "__main__":
    result = main()