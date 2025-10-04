import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def clean_company_name(name):
    """Standardize company name for matching"""
    if pd.isna(name):
        return ""
    
    name = str(name).upper().strip()
    
    # Remove common suffixes
    suffixes = [
        r'\s+LLC\.?$', r'\s+L\.L\.C\.?$', r'\s+L\s*L\s*C$',
        r'\s+INC\.?$', r'\s+INCORPORATED$',
        r'\s+CORP\.?$', r'\s+CORPORATION$',
        r'\s+PA$', r'\s+P\.A\.?$',
        r'\s+LTD\.?$', r'\s+LIMITED$',
        r'\s+CO\.?$', r'\s+COMPANY$',
        r'\s+PLC$', r'\s+LP$', r'\s+L\.P\.?$'
    ]
    
    for suffix in suffixes:
        name = re.sub(suffix, '', name)
    
    # Remove punctuation and extra spaces
    name = re.sub(r'[^\w\s]', ' ', name)
    name = re.sub(r'\s+', ' ', name)
    
    return name.strip()

def extract_address_info(line):
    """Extract address from the officer line data"""
    # Address typically starts around position 750-850
    # Format: ADDRESS + spaces + CITY + spaces + STATE + ZIP
    
    address_section = line[750:].strip() if len(line) > 750 else ""
    
    # Try to parse: ADDRESS, CITY, STATE ZIP
    parts = address_section.split()
    
    if len(parts) < 3:
        return "", ""
    
    # Simple heuristic: Last 2 items are usually STATE ZIP
    # Everything before is ADDRESS + CITY
    
    # Look for FL followed by 5 digits
    state_zip_pattern = r'(FL|FLORIDA)\s*(\d{5}[-\d]*)'
    match = re.search(state_zip_pattern, address_section)
    
    if match:
        state = match.group(1)
        zip_code = match.group(2)
        
        # Everything before the state is address + city
        before_state = address_section[:match.start()].strip()
        
        # Split into address and city (last word before state is usually city)
        parts = before_state.split()
        if len(parts) > 1:
            city = parts[-1]
            address = ' '.join(parts[:-1])
        else:
            city = before_state
            address = ""
        
        city_state_zip = f"{city}, {state} {zip_code}"
        return address, city_state_zip
    
    return "", ""

def match_and_fill(corps_excel, officers_csv):
    """Match companies and fill in officer data with formatting"""
    
    print("Loading data...")
    print("="*60)
    
    # Load company list
    companies_df = pd.read_excel(corps_excel)
    print(f"Companies to match: {len(companies_df):,}")
    
    # Load officers data
    officers_df = pd.read_csv(officers_csv)
    print(f"Officer records available: {len(officers_df):,}")
    
    # Clean company names for matching
    companies_df['clean_name'] = companies_df['Company'].apply(clean_company_name)
    officers_df['clean_name'] = officers_df['company_name'].apply(clean_company_name)
    
    # Get first officer per company (sorted by line_number to get the first listed)
    print("\nPreparing officer data (first officer per company)...")
    officers_first = officers_df.sort_values('line_number').groupby('clean_name').first().reset_index()
    print(f"Unique companies with officers: {len(officers_first):,}")
    
    # Merge
    print("\nMatching companies...")
    merged = companies_df.merge(
        officers_first[['clean_name', 'officer_full_name', 'officer_role', 'status', 'line_number']],
        on='clean_name',
        how='left'
    )
    
    # Set default address info (since we don't have source_line data)
    print("Setting default address information...")
    merged['address'] = ""
    merged['city_state_zip'] = ""
    
    # Create output DataFrame with correct columns
    result = pd.DataFrame({
        'Company': companies_df['Company'],
        'Officer': merged['officer_full_name'],
        'Address': merged['address'],
        'City State Zip': merged['city_state_zip']
    })
    
    # Statistics
    matched = result['Officer'].notna().sum()
    print(f"\n{'='*60}")
    print(f"RESULTS:")
    print(f"{'='*60}")
    print(f"Total companies: {len(result):,}")
    print(f"Matched: {matched:,} ({matched/len(result)*100:.1f}%)")
    print(f"Unmatched: {len(result) - matched:,}")
    
    # Save with formatting
    output_file = 'Corps_with_Officers_Matched.xlsx'
    print(f"\nSaving to {output_file}...")
    
    result.to_excel(output_file, index=False)
    
    # Apply blue fill to all cells
    print("Applying formatting...")
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Light blue fill (matching your screenshot)
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    
    # Apply to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.fill = blue_fill
    
    wb.save(output_file)
    
    print(f"\n✓ Complete! Saved to: {output_file}")
    
    # Show sample of unmatched
    unmatched = result[result['Officer'].isna()]
    if len(unmatched) > 0:
        print(f"\nSample of unmatched companies (first 10):")
        for company in unmatched['Company'].head(10):
            print(f"  • {company}")
    
    return result

# Run the matching
result = match_and_fill('Corps 10-2-25.xlsx', 'officers_clean.csv')